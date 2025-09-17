import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys
import json
from datetime import datetime
import re

def process_lightware_excel(input_file_path, template_path, output_path):
    """
    Process Lightware Excel file and populate the Q360 template.
    
    Mappings:
    - Part number → MASTERNO
    - Product name → PARTNO  
    - Description → DESCRIPTION
    - PSNI PARTNER COST → STANDARDCOST
    - MSRP USD → MSRP
    """
    try:
        # Load the source file - header is at row 2 (index 2)
        df_source = pd.read_excel(input_file_path, header=2)
        
        # Clean column names (remove any extra spaces)
        df_source.columns = df_source.columns.str.strip()
        
        # Check if expected columns exist
        required_columns = ['Part number', 'Product name', 'Description', 'MSRP USD', 'PSNI PARTNER COST']
        missing_columns = [col for col in required_columns if col not in df_source.columns]
        
        if missing_columns:
            return {"error": f"Missing columns in source file: {missing_columns}. Found columns: {list(df_source.columns)}"}
        
        # Filter out rows where PSNI PARTNER COST is not a number
        df_filtered = df_source.copy()
        
        # Convert PSNI PARTNER COST to numeric, invalid parsing will result in NaN
        df_filtered['PSNI PARTNER COST'] = pd.to_numeric(df_filtered['PSNI PARTNER COST'], errors='coerce')
        
        # Remove rows where PSNI PARTNER COST is NaN (was text)
        df_filtered = df_filtered.dropna(subset=['PSNI PARTNER COST'])
        
        # Also remove rows where any key field is empty
        df_filtered = df_filtered.dropna(subset=['Part number', 'Product name', 'Description', 'MSRP USD'])
        
        # Load the template
        wb = load_workbook(template_path)
        sheet = wb.active
        
        # Clear existing data (keep headers)
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(2, max_row - 1)
        
        # Map and populate data
        for idx, row in df_filtered.iterrows():
            excel_row = idx - df_filtered.index[0] + 2  # Start from row 2 in Excel
            
            # Mapping as per requirements:
            # Part number → MASTERNO (column A)
            sheet[f'A{excel_row}'] = str(row['Part number'])
            
            # Product name → PARTNO (column C) 
            sheet[f'C{excel_row}'] = str(row['Product name'])
            
            # Description → DESCRIPTION (column E)
            sheet[f'E{excel_row}'] = str(row['Description'])
            
            # PSNI PARTNER COST → STANDARDCOST (column S)
            sheet[f'S{excel_row}'] = float(row['PSNI PARTNER COST'])
            
            # MSRP USD → MSRP (column T)
            sheet[f'T{excel_row}'] = float(row['MSRP USD'])
            
            # Set brand as Lightware in MANUFACTURER (column I)
            sheet[f'I{excel_row}'] = 'Lightware'
            
            # Set TAXABLE and USETAXFLAG to Y (columns J and K)
            sheet[f'J{excel_row}'] = 'Y'
            sheet[f'K{excel_row}'] = 'Y'
        
        # Save the workbook
        wb.save(output_path)
        
        return {
            "success": True, 
            "rows_processed": len(df_filtered),
            "rows_excluded": len(df_source) - len(df_filtered),
            "output_file": output_path
        }
        
    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python excel_processor.py <input_file> <template_file> <output_file>")
        sys.exit(1)
    
    result = process_lightware_excel(sys.argv[1], sys.argv[2], sys.argv[3])
    print(json.dumps(result))
